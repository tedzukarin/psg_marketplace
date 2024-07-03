import datetime
import openpyxl.worksheet.worksheet
import pandas as pd
from openpyxl.styles import (PatternFill, Border, Side)
from openpyxl import load_workbook
import re
import PySimpleGUI as sg
import PyPDF2 as pdf
import os.path
import json
import string
import time
import ostatki_func
import prod_func
import analitics
import api
import change_price
import change_size_wb
import image_change
from reportlab.pdfgen import canvas
from PIL import Image
from reportlab.lib.utils import ImageReader


version = 'Market_Places_0.961'
server = '82.146.35.1'
port = '5555'
to_day = datetime.date.today()

def ozon_prodazhi_na_ukazannyi_den(detal_name, day):
    if day == '':
        sg.Print('Не выбрана дата отгрузки')
        return
    day_time = day + ' 10:00:00'
    zakaz_csv = pd.read_csv(f'{detal_name}', delimiter=';', dtype=str, quotechar='"')
    zakaz_csv['Артикул'] = zakaz_csv['Артикул'].astype(str)
    zakaz_csv_new = zakaz_csv[zakaz_csv['Дата отгрузки'] == day_time]
    if len(zakaz_csv_new) == 1:
        sg.Print('Не нашлось отфильтрованных данных')
        return
    zakaz_csv_new.to_csv(f'{detal_name[:-4]}_{day}.csv', encoding='utf-8', sep=';', index=False)
    sg.Print('Отправления за выбранный день отфильтрованны')

def ozon_ostatki(ozon_out_name, zakaz_csv_name, name_last_file = False):
    '''
    Формирует лист подбора, возвращает при этом список номеров отправления (для дальнейшей сортировки ярлыков)
    :param ozon_out_name: название выходного файла
    :param zakaz_csv_name: название входного файла
    :return: список отправлений
    '''
    settings_file = open('settings.ini', 'r')
    settings = json.load(settings_file)
    ozon_number_column = settings['ozon_number_column']
    ozon_kol_column = settings['ozon_kol_column']
    ozon_name_column = settings['ozon_name_column']
    ozon_article_column = settings['ozon_article_column']
    other_columns = settings['ozon_other_columns']
    color_1 = settings['ozon_color_1']
    color_2 = settings['ozon_color_2']
    ozon_special_delimeter = settings['ozon_special_delimeter']
    ozon_column_dimension = settings['ozon_column_dimension']
    ozon_column_dimension = ozon_column_dimension.split(';')


    if len(other_columns) > 1:
        other_columns = other_columns.split(';')

    zakaz_csv = pd.read_csv(f'{zakaz_csv_name}', delimiter=';', encoding='utf-8')
    number_of = zakaz_csv[f'{ozon_number_column}'].tolist()
    current_number = ''
    n = -1
    index_zak = []
    for i in number_of:
        if current_number == i:
            index_zak.append(n)
        else:
            current_number = i
            n += 1
            index_zak.append(n)

    zakaz_csv['Индекс'] = index_zak
    need_columns = [f'{ozon_number_column}', f'{ozon_name_column}', f'{ozon_article_column}', f'{ozon_kol_column}',
                    *other_columns, 'Индекс']

    '''Формируем новую таблицу. Оставляем только нужные столбцы'''
    new_zakaz = zakaz_csv[need_columns]

    if name_last_file:
        last_table = pd.read_csv(f'{name_last_file}', delimiter=';')
        last_numbers = last_table[f'{ozon_number_column}'].tolist()
        current_number = new_zakaz[ozon_number_column].tolist()
        new_zakaz = new_zakaz[~new_zakaz[ozon_number_column].isin(last_numbers)]
        delete_orders = []
        for i in last_numbers:
            if i not in current_number:
                delete_orders.append(i)
        sg.Print(f'Эти номера отправлений не были найдены в последнем файле:\n{delete_orders}')

    '''Отбираем сборные задания'''
    grupp_zakaz = new_zakaz[new_zakaz.duplicated([f'{ozon_number_column}'], keep=False)]

    '''Фиксируем и сортируем оставшиеся единичные заказы'''
    solo_zakaz = new_zakaz.drop_duplicates(f'{ozon_number_column}', keep=False)
    solo_zakaz = solo_zakaz.sort_values([f'{ozon_name_column}', f'{ozon_kol_column}'], ascending=[True, False])


    new_group_zakaz = pd.DataFrame(columns=need_columns) # Создали новый дата-фрейм
    null_serries = pd.Series({need_columns[0]: None, need_columns[1]: None, need_columns[2]: None, need_columns[3]: None},
                             index=None) # пустая строка для разделения
    null_row = ''

    '''Вставляем пустую разделительную строку при разбивке на сборки'''
    for _, row in grupp_zakaz.iterrows():
        if null_row == '':
            null_row = row[0]
            new_group_zakaz = new_group_zakaz._append(row, ignore_index=True)
        elif row[0] != null_row:
            null_row = row[0]
            new_group_zakaz = new_group_zakaz._append(null_serries, ignore_index=True)
            new_group_zakaz = new_group_zakaz._append(row, ignore_index=True)
        else:
            new_group_zakaz = new_group_zakaz._append(row, ignore_index=True)
    new_group_zakaz = new_group_zakaz._append(null_serries, ignore_index=True)

    '''Объединяем сборные заказы с индивидуальными'''
    new_group_zakaz = new_group_zakaz._append(solo_zakaz)

    '''Собираем все номера отправки для дальнейшей обработки ярлыков'''
    numbers_of_orders = new_group_zakaz[need_columns[0]].tolist()
    index_of_orders = new_group_zakaz["Индекс"].tolist()

    '''Сохраняем новую книгу'''
    new_group_zakaz[need_columns[:-1]].to_excel(ozon_out_name, index = False)

    '''Открываем книгу для внесения стилистических изменений'''
    wb = load_workbook(ozon_out_name)
    ws = wb.active

    kol_list = []
    fas_list = []

    '''Отбираем индексы строк с фасовкой, отличной от стандарта и с количеством товара > 1'''
    for ind, val in enumerate(ws.values):
        if val[3] and isinstance(val[3], int) and val[3] > 1:
            kol_list.append(ind)
        if ozon_special_delimeter:
            if val[2] and ozon_special_delimeter in val[2]:
                fas_list.append(ind)

    color_kol = PatternFill(fill_type='solid', fgColor=color_1)
    color_fas = PatternFill(fill_type='solid', fgColor=color_2)

    literal = list(string.ascii_uppercase)
    for i in range(len(ozon_column_dimension)):
        ws.column_dimensions[literal[i]].width = ozon_column_dimension[i]


    for ind, row in enumerate(ws.rows):
        for i in kol_list:
            if ind == i:
                for c in row:
                    c.fill = color_kol
        for j in fas_list:
            if ind == j:
                for c in row:
                    c.fill = color_fas

    '''Добавляем границы к ячейкам'''
    thins = Side(border_style="thin", color="000000")
    for row in ws:
        for cell in row:
            cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

    '''Меняем ориентацию страницы на альбомную'''
    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, 9, orientation='landscape')

    '''Сохраняем выходной файл'''
    wb.save(f'{ozon_out_name}')
    return numbers_of_orders, index_of_orders

def sorting_labels_ozon(labels_file_name, out_file_name, numbers_of_orders, index_of_orders, flag):
    '''Создаём новый pdf обработчик для нового файла'''
    pdf_writer = pdf.PdfWriter()

    '''Открываем и создаём объект обработчика'''
    pdf_file = open(f'{labels_file_name}', 'rb')
    pdf_reader = pdf.PdfReader(pdf_file)

    '''Получили количество страниц'''
    num_pages = pdf_reader.pages
    if flag:
        number_set = set()  # Пустое множестводля проверки уже пройденных кодов
        page_set = set()
        pages_list = [i for i in range(len(num_pages))]  # необработанных страниц

        for i in numbers_of_orders:  # для каждого номера отправки:
            if i not in page_set:
                for page in pages_list:  # Для каждого номера страницы:
                    if page not in number_set:
                        page_obj = pdf_reader.pages[page]  # Получили объект страницы
                        text = page_obj.extract_text()  # Получили текст
                        text = text.split('\n')  # Разбили по строкам
                        numbers_of_order = text[2].split(' ')[0]  # Выбрали номер отправлния
                        if i == numbers_of_order:
                            pdf_writer.add_page(page_obj)
                            pages_list.remove(page)
                            number_set.add(page)
                            sg.Print(f'Осталось необработанных ярлыков OZON: {len(pages_list)}')
                            page_set.add(i)
                            break

    else:
        current = ''
        for i in index_of_orders:
            if pd.isna(i) or i == current:
                continue
            else:
                current = i
                page_obj = pdf_reader.pages[int(i)]
                pdf_writer.add_page(page_obj)
    day = to_day
    if values['data_for_check_csv']:
        day = values['data_for_check_csv']


    if len(pdf_writer.pages) != len(num_pages):
        sg.Print('Длина выходного списка не соответствует входному')
        with open(f'{out_file_name}_{day}.pdf','wb') as out:
            pdf_writer.write(out)
    else:
        with open(f'{out_file_name}_{day}.pdf','wb') as out:
            pdf_writer.write(out)
        sg.Print('Файл ярлыков изменён')
    # Закрываем файл
    pdf_file.close()

def get_numbers_of_labels(podbor_file_name, wb_first_part_number, wb_second_part_number):
    '''Открываем и создаём объект обработчика'''
    pdf_file = open(f'{podbor_file_name}', 'rb')
    pdf_reader = pdf.PdfReader(pdf_file)
    wb_first_part_number = '{' + wb_first_part_number + '}'
    wb_second_part_number = '{' + wb_second_part_number + '}'
    '''Получили количество страниц'''
    num_pages = pdf_reader.pages
    pages_list = [i for i in range(len(num_pages))]  # необработанных страниц
    numbers_of_labels = []
    for page in pages_list:
        page_obj = pdf_reader.pages[page]  # Получили объект страницы
        text = page_obj.extract_text()
        text = text.split('\n')  # Получили текст

        '''шаблон для прямого поиска. Сейчас используется инвертирование текста для того, что поиск бежал с конца строки
        Так работает чуть дольше, зато исключет случаи, когда к артикулу добавляются неприемлимые символы из-за переноса
        строки. 
        Что бы запустить эту версию, нужно вернуть проверку i, а не s, а так же для number_of_label убрать 
        инвертирование'''
        # pattern = f'[ а-яА-Яa-zA-Z]\d{wb_first_part_number} \d{wb_second_part_number}

        pattern = f'\d{wb_second_part_number} \d{wb_first_part_number}'
        for i in text:  # Вычленяем их
            s = i[::-1]
            match = re.findall(pattern, s)
            if len(match) != 0:
                number_of_label = str(*match).replace(' ', '')[::-1]
                numbers_of_labels.append(number_of_label)



    pdf_file.close()
    return numbers_of_labels

def sorting_labels_wb(labels_file_name, out_file_name, numbers_of_orders):
    '''Создаём новый pdf обработчик для нового файла'''
    pdf_writer = pdf.PdfWriter()

    '''Открываем и создаём объект обработчика'''
    pdf_file = open(f'{labels_file_name}', 'rb')
    pdf_reader = pdf.PdfReader(pdf_file)

    '''Получили количество страниц'''
    num_pages = pdf_reader.pages
    pages_list = [i for i in range(len(num_pages))]  # необработанных страниц
    page_obj_and_numbers_list = {}

    '''версия сортировки на основании всего списка. Работает быстрее'''
    for page in pages_list:
        page_obj = pdf_reader.pages[page]
        text = page_obj.extract_text()
        text = text.replace('\n', '')[2:]
        page_obj_and_numbers_list[text] = page_obj
    for i in numbers_of_orders:
        if i in page_obj_and_numbers_list.keys():
            pdf_writer.add_page(page_obj_and_numbers_list[i])
        else:
            sg.Print('Не найден номер отправления: ', i)
            return


    '''Версия с принудительным поиском и сравнением ярлыков. Работает медленнее. Выводит на экран текущий прогресс'''
    # number_set = set()  # Пустое множестводля проверки уже пройденных кодов
    # for i in numbers_of_orders:  # для каждого номера отправки:
    #     for page in pages_list:  # Для каждого номера страницы:
    #         if page not in number_set:
    #             page_obj = pdf_reader.pages[page]  # Получили объект страницы
    #             text = page_obj.extract_text()  # Получили текст
    #             text = text.replace('\n', '')  # сшили строки
    #             if i == text:
    #                 pdf_writer.add_page(page_obj)
    #                 pages_list.remove(page)
    #                 number_set.add(page)
    #                 sg.Print(f'{i} Осталось необработанных ярлыков WB: {len(pages_list)}')
    #                 break


    if len(pdf_writer.pages) != len(num_pages):
        sg.Print('Длина выходного списка не соответствует входному')

    else:
        with open(f'{out_file_name}','wb') as out:
            pdf_writer.write(out)
        sg.Print(f'{out_file_name} изменён')

    # Закрываем файл
    pdf_file.close()

def compile_wb(list_of_files, file_name = 'Compile', change_price_flag=False):
    for ind, i in enumerate(list_of_files):
        current_df = pd.read_excel(i, converters={'Размер': str, 'Артикул продавца': str})
        if ind == 0:
            big_df = current_df
            razmer = big_df.shape[1]
        else:
            if razmer != current_df.shape[1]:
                sg.Print(f'Количество столбцов в таблице {i} не совпадает с первой таблицей')
                sg.Print('Проверьте все вводимые таблицы и попробуйте снова')
                return
            big_df = pd.concat([big_df, current_df], ignore_index=True)
    dirname, filename = os.path.split(list_of_files[0])

    if change_price_flag:
        sg.Print('Установлен флаг замены цены')
        sg.Print('Получаем цены')
        new_price = api.get_price_wb()
        sg.Print('Цены получены')

        new_price = {i['nmID']: i['sizes'][0]['discountedPrice'] for i in new_price}
        def apply_to(row):
            if row['Артикул Wildberries'] in new_price.keys():
                return new_price[row['Артикул Wildberries']]
        big_df['Стоимость'] = big_df.apply(apply_to, axis=1)

    big_df.to_excel(dirname + f'\{file_name}.xlsx', index=False)
    sg.Print(f'Файл успешно объеденён и сохранён в каталог:\n{dirname}')

def crop_pdf(pdf_file):
    # Открываем PDF файл
    pdf_ = open(f'{pdf_file}', 'rb')
    pdf_reader = pdf.PdfReader(pdf_)

    # Создаем новый PDF файл
    pdf_writer = pdf.PdfWriter()

    # Получаем размер страницы A4
    page_width = pdf_reader.pages[0].mediabox[2]
    page_height = pdf_reader.pages[0].mediabox[3]

    # Разделяем страницу на 6 частей
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        page.mediabox.lowe_left = (12, 12)
        page.mediabox.lower_right = (page_width - 12, 12)
        page.mediabox.upper_left = (12, page_height - 12)
        page.mediabox.upper_right = (page_width - 12, page_height - 12)
        _width = page.mediabox[2]
        _height = page.mediabox[3]

        for i in range(3):
            for j in range(2):
                page.cropbox.lower_left = (12 + j * _width / 2, (3 - i) * _height / 3)
                page.cropbox.upper_right = ((j + 1) * _width / 2, 13 + (2 - i) * _height / 3)
                pdf_writer.add_page(page)
    pdf_.close()

    # Сохраняем результат в новый файл
    with open(f'{pdf_file}', 'wb') as output_file:
        pdf_writer.write(output_file)

    # # Не получается найти решение очистки пустых страниц. Потому что каждая имеет текстовый невидимый слой
    # pdf_ = open(f'{pdf_file}_', 'rb')
    # new_pdf_reader = pdf.PdfReader(pdf_)
    # pdf_writer = pdf.PdfWriter()
    # for page_num in range(len(new_pdf_reader.pages)):
    #     new_page = new_pdf_reader.pages[page_num]
    #     pdf_writer.add_page(new_page)
    #
    # pdf_.close()
    # # Сохраняем результат в новый файл
    # with open(f'{pdf_file}_', 'wb') as output_file:
    #     pdf_writer.write(output_file)


def check_time(settings, use_data):
    cur_time = datetime.datetime.now().time()
    cur_date = datetime.datetime.now().date()
    sep_set = {',', '/', '.', ' ', ':', ';'}

    if settings['first_time']:
        if settings['first_time'].isdigit():
            utro = datetime.time(int(settings['first_time']))
        else:
            flag = False
            for i in sep_set:
                if i in settings['first_time']:
                    sep = i
                    flag = True
            if not flag:
                sg.Print('Разделитель в настройках времени указан не верно. Настройки времени сброшены на 0.0.0\n'
                                'Используйте любой из предложенных (.,;:/ или пробел)')
                utro = datetime.time(0)
            else:
                utro = datetime.time(*[int(i) for i in settings['first_time'].split(sep)])
    else:
        utro = datetime.time(0)

    if settings['second_time']:
        if settings['second_time'].isdigit():
            vecher = datetime.time(int(settings['second_time']))
        else:
            flag = False
            for i in sep_set:
                if i in settings['second_time']:
                    sep = i
                    flag = True
            if not flag:
                sg.Print('Разделитель в настройках времени указан не верно. Настройки времени сброшены на 0.0.0\n'
                         'Используйте любой из предложенных (.,;:/ или пробел)')
                vecher = datetime.time(0)
            else:
                vecher = datetime.time(*[int(i) for i in settings['second_time'].split(sep)])
    else:
        vecher = datetime.time(0)

    if cur_time < utro:
        compile_text = settings['first_time_name']
    elif utro < cur_time < vecher:
        compile_text = settings['second_time_name']
    elif cur_time > vecher:
        compile_text = settings['third_time_name']

    if use_data:
        compile_text += cur_date.strftime("%Y%m%d")
    if len(compile_text) == 0:
        compile_text = 'Pass'
    return compile_text

if __name__ == '__main__':
    settings = open('settings.ini', 'r')
    settings = json.load(settings)

    try:
        compile_text = check_time(settings, settings['use_datetime'])
    except Exception as P:
        compile_text = 'Pass'
        sg.Print(P)
        sg.Print('Нарушен формат времени. Советуем использовать ЧЧ.ММ.СС\n'
                 'Не ставьте пробелы. Если точность до секунд или минут не нужна - указывайте только часы или часы с минутами')

    sg.theme("DarkGray6")
    layout = [[sg.T("")], [sg.Text("Выберите лист подбора: "), sg.Input(),
                           sg.FilesBrowse(key="podbor", button_text='Выбрать', file_types=(('CSV, PDF', ('*.pdf', '*.csv')),))],
              [sg.Text("Прошлая детализация:   "), sg.Input(), sg.FileBrowse(key="last_order", button_text='Выбрать',
                                                                              enable_events=True, disabled=True)],
              [sg.Text("Выберите ярлыки:          "), sg.Input(), sg.FileBrowse(key="label", button_text='Выбрать',
                                                                                file_types=(
                                                                                ('PDF', ('*.pdf')),))],
              [sg.Input(size=(12, 1), key='data_for_check_csv'), sg.CalendarButton('Дата отгрузки ОЗОН', format='%Y-%m-%d'),
               sg.Checkbox('Сортировка всего и сразу', key='wb_sorting_all')],
              [sg.Button('Сформировать CSV ОЗОН для выбранной даты отгрузки')],
              [sg.Text("Укажите Площадку: ")],
              [sg.Radio('WB', "RADIO1", default=True, key='wb', enable_events=True),
               sg.Radio('SBER', "RADIO1", default=False, key='sber', enable_events=True),
               sg.Radio('OZON', "RADIO1", default=False, key='ozon', enable_events=True),
               sg.Checkbox('Без учёта предв. сорт', default=False, key='pred_sort', disabled=True, enable_events=True)],
              [sg.T("")], [sg.Text("Файлы детализации:      "), sg.Input(),
                           sg.FilesBrowse(key="compile", button_text='Выбрать', file_types=(('Excell', '*.xlsx'),))],
              [sg.Radio('Объединить детализации WB', "RADIO1", default=False, key='compile_wb'), sg.Text("      Имя файла:"),
               sg.Input(key='name_compile', size=22, default_text=f'{compile_text}'
    if settings["use_auto"] == True else settings["default_name"], enable_events=True)],
              [sg.Checkbox('Изменить цены для WB в соотвтетствии с актуальными ценами', default=False, key='check_price_wb')],
              [sg.T("")],
              [sg.Text("Файл остатков:              "), sg.Input(),
               sg.FilesBrowse(key="ostat", button_text='Выбрать', file_types=(('Excell', '*.xlsx'),))],
              [sg.Text("Файл резерва:              "), sg.Input(),
               sg.FilesBrowse(key="rezerv", button_text='Выбрать', file_types=(('Excell', '*.xlsx'),))],
              [sg.Text("Файл продаж  :             "), sg.Input(),
               sg.FilesBrowse(key="prod", button_text='Выбрать', file_types=(('Excell', '*.xlsx'),))],
              [sg.Radio('Распределить остатки', "RADIO1", default=False, key='make_ostat'),
               sg.Checkbox('Только сбер', key='sber_flag'), sg.Checkbox('Обновить по API', key='use_api_flag')],
               [sg.Radio('Проверить остатки', "RADIO1", default=False, key='make_ost_prod'),
              sg.Text("      Имя файла "), sg.Input(key="prod_name", size=17)],
              [sg.Button("Начать"), sg.T(""), sg.Button("Анализ заказов"), sg.T(""), sg.Button("Работа с заказами"),
               sg.T(""), sg.Button("Изменить цены"), sg.T(""), sg.Button("Изм. габариты WB")],
              [sg.T("")], [sg.T("")], [sg.Button("Настройка сортировки"), sg.T("    "), sg.Button("Настройка остатков"),
                                       sg.T("    "), sg.Button("Настройка проверки остатков")]]

    ###Building Window
    window = sg.Window(version, layout, resizable=True, size=(650, 700))
    # try:
    #     token = open('token.txt').read()
    # except:
    #     sg.Print('Не удаётся найти файл "token.txt"')
    #     time.sleep(5)
    # try:
    #     response = requests.get(f'http://{server}:{port}/?token={token}')
    #
    #     if response.text == 'False':
    #         sg.Print('Время действия токена истекло')
    #         time.sleep(5)
    #
    #     elif response.text == 'True':
    while True:
        elem = window['pred_sort']
        last_order = window['last_order']
        name_elem = window['name_compile']
        event, values = window.read()
        settings = open('settings.ini', 'r')
        settings = json.load(settings)
        ostatki_set = open('settings_ost.ini', 'r')
        ostatki_set = json.load(ostatki_set)
        prod_set = open('settings_ost_prod.ini', 'r')
        prod_set = json.load(prod_set)

        if event == sg.WIN_CLOSED or event == "Exit":
            break
        elif event == 'Сформировать CSV ОЗОН для выбранной даты отгрузки':
            try:
                ozon_prodazhi_na_ukazannyi_den(detal_name=values['podbor'], day=values['data_for_check_csv'])
            except Exception as P:
                sg.Print(P)
        elif event == "Начать":
            if values['wb']:
                if values['wb_sorting_all']:
                    try:
                        label_orders_dict = {}
                        for i in values['podbor'].split(';'):
                            word_ticket = settings['word_ticket']
                            word_order = settings['word_order']
                            if word_ticket in i:
                                podbor_name = i.replace(word_ticket, word_order)
                                label_orders_dict[i] = podbor_name
                                tic = time.perf_counter()
                                numbers_of_labels = get_numbers_of_labels(podbor_name,
                                                                          wb_first_part_number=settings[
                                                                              'wb_first_part_number'],
                                                                          wb_second_part_number=settings[
                                                                              'wb_second_part_number'])
                                sorting_labels_wb(i, i, numbers_of_labels)
                                toc = time.perf_counter()
                                sg.Print(f'Время обработки: {round(toc - tic, 2)}')
                        sg.Print('Завершено')
                    except Exception as P:
                        sg.Print(P)
                else:
                    try:
                        tic = time.perf_counter()
                        numbers_of_labels = get_numbers_of_labels(values['podbor'],
                                                                  wb_first_part_number=settings['wb_first_part_number'],
                                                                  wb_second_part_number=settings['wb_second_part_number'])
                        sorting_labels_wb(values['label'], values['label'], numbers_of_labels)
                        toc = time.perf_counter()
                        sg.Print(f'Время обработки: {round(toc-tic, 2)}')
                    except Exception as P:
                        sg.Print(P)
            elif values['sber']:
                sber_file = values['label']
                crop_pdf(sber_file)
                sg.Print('Файл поделён на ярлыки')
            elif values['ozon']:
                try:
                    tic = time.perf_counter()
                    numbers_of_orders, index_zak = ozon_ostatki(f'{values["podbor"][:-4]}.xlsx', values['podbor'], name_last_file=values['last_order'])
                    sorting_labels_ozon(values['label'], values['label'][:-4], numbers_of_orders, index_zak, flag=values['pred_sort'])

                    toc = time.perf_counter()
                    sg.Print(f'Время обработки: {round(toc-tic, 2)}')
                except Exception as P:
                    sg.Print(P)

            elif values['compile_wb']:
                try:
                    compile_wb(values['compile'].split(';'), values['name_compile'], values['check_price_wb'])
                except Exception as P:
                    sg.Print(P)

            elif values['make_ostat']:
                try:
                    ostatki_func.raspr_ost(ostat=values['ostat'], rezerv=values['rezerv'], sber_flag=values['sber_flag'],
                                           use_api_flag=values['use_api_flag'])
                except Exception as P:
                    sg.Print(P)

            elif values['make_ost_prod']:
                try:
                    sg.Print('Начинаем')
                    prod, days = prod_func.get_prod(values['prod'])
                    sg.Print('Продолжаем')
                    ost = prod_func.get_ost(values['ostat'],art=ostatki_set['art_kol'],
                                            kol= ostatki_set['kol_kol'], skip_row=ostatki_set['skip_row'], )
                    sg.Print('Ещё')
                    prod_func.make_tab(prod, ost, days, prod_set['prod_out_path']+'/'+values['prod_name'])
                    sg.Print('Закончили')
                except Exception as P:
                    sg.Print(P)

        elif event == "Настройка сортировки":
            layout_setting = [[sg.Text('Для указания диапазона используйте формат: "n,m"\n'
                                       'Где n - начало, а m - конец (включая его). Без пробела')],
                              [sg.Text('Ожидаемое количество цифр в первой части номера отправлений: '),
                               sg.Input(key='wb_first_part_number', size=3,
                                        default_text=f'{settings["wb_first_part_number"]}')],
                               [sg.Text('Ожидаемое количество цифр во второй части номера отправлений: '),
                               sg.Input(key='wb_second_part_number', size=3,
                                       default_text=f'{settings["wb_second_part_number"]}')],
                              [sg.T("")],
                              [sg.Text('Настройки WB')],
                              [sg.Text('Слово для ярлыков:     '), sg.Input(key='word_ticket', size=56,
                                        default_text=f'{settings["word_ticket"]}')],
                              [sg.Text('Слово для листов подбора:     '), sg.Input(key='word_order', size=56,
                                                                            default_text=f'{settings["word_order"]}')],

                              [sg.Text('Настройки OZON')],
                              [sg.Text('Колонка с номерами отправлений:     '),
                               sg.Input(key='ozon_number_column', size=56,
                                        default_text=f'{settings["ozon_number_column"]}')],
                              [sg.Text('Колонка с количеством отправлений: '),
                               sg.Input(key='ozon_kol_column', size=56,
                                        default_text=f'{settings["ozon_kol_column"]}')],
                              [sg.Text('Колонка с наименованием товаров:   '),
                               sg.Input(key='ozon_name_column', size=56,
                                        default_text=f'{settings["ozon_name_column"]}')],
                              [sg.Text('Колонка с артикулом:                        '),
                               sg.Input(key='ozon_article_column', size=56,
                                        default_text=f'{settings["ozon_article_column"]}')],
                              [sg.Text('Дополнительные колонки через ; :     '),
                              sg.Input(key='ozon_other_columns', size=56,
                                       default_text=f'{settings["ozon_other_columns"]}')],
                              [sg.Text('Цвет количества: '),
                               sg.Input(key='ozon_color_1', size=10,
                                        default_text=f'{settings["ozon_color_1"]}'),
                              sg.Text('   Цвет специального разделителя: '),
                               sg.Input(key='ozon_color_2', size=10,
                                        default_text=f'{settings["ozon_color_2"]}')],
                              [sg.Text('Специальный разделитель: '),
                               sg.Input(key='ozon_special_delimeter', size=10,
                                        default_text=f'{settings["ozon_special_delimeter"]}'),
                              sg.Text('   Размеры колонок через ; :'),
                               sg.Input(key='ozon_column_dimension', size=26,
                                        default_text=f'{settings["ozon_column_dimension"]}')],
                              [sg.T("")],
                              [sg.Text('Настройка объединения детализаций')],
                              [sg.Checkbox('Использовать автозаполнение', default=settings['use_auto'],
                                           key='use_auto', enable_events=True),
                               sg.Checkbox('Иcпользовать дату', default=settings['use_datetime'],
                                 key='use_datetime', enable_events=True)],
                              [sg.Text('до'),
                               sg.Input(key='first_time', default_text=settings['first_time'], enable_events=True,
                                        size=10),
                               sg.Text(' часов –'),
                               sg.Input(key='first_time_name', default_text=settings['first_time_name'],
                                        enable_events=True, size=20)],
                              [sg.Text('до'),
                               sg.Input(key='second_time', default_text=settings['second_time'], enable_events=True,
                                        size=10),
                               sg.Text(' часов –'),
                               sg.Input(key='second_time_name', default_text=settings['second_time_name'],
                                        enable_events=True, size=20),
                              sg.Text('а после'),
                               sg.Input(key='third_time_name', default_text=settings['third_time_name'],
                                        enable_events=True, size=20)],

                              [sg.Text('Имя объединённого файла по умолчанию: '), sg.Input(
                                  default_text=settings['default_name'], key='default_name', enable_events=True,
                              disabled=settings['use_auto'])],
                              [sg.Text('Общие характеристики одним файлом'), sg.Input(size=35, key='all_attributes',
                                                                                      default_text=settings['all_attributes']),
                               sg.FilesBrowse('Выбрать')],
                              [sg.T("")], [sg.Button("Сохранить")]
                              ]

            window_setting = sg.Window('Настройка сортировки', layout_setting, resizable=True, size=(620, 680))
            while True:
                elem_default_name = window_setting['default_name']
                elem_list = [window_setting['first_time'], window_setting['first_time_name'],
                             window_setting['second_time'], window_setting['second_time_name'],
                             window_setting['third_time_name'], window_setting['use_datetime']]
                event_set, values_set = window_setting.read()
                if event_set == sg.WIN_CLOSED or event_set == "Exit":
                    break

                elif event_set == "Сохранить":
                    setting_set = {
                        'wb_first_part_number': values_set['wb_first_part_number'],
                        'wb_second_part_number': values_set['wb_second_part_number'],
                        'ozon_number_column': values_set['ozon_number_column'],
                        'ozon_kol_column': values_set['ozon_kol_column'],
                        'ozon_name_column': values_set['ozon_name_column'],
                        'ozon_article_column': values_set['ozon_article_column'],
                        'ozon_other_columns': values_set['ozon_other_columns'],
                        'ozon_color_1': values_set['ozon_color_1'],
                        'ozon_color_2': values_set['ozon_color_2'],
                        'ozon_special_delimeter': values_set['ozon_special_delimeter'],
                        'ozon_column_dimension': values_set['ozon_column_dimension'],
                        'use_datetime': values_set['use_datetime'],
                        'default_name': values_set['default_name'],
                        'first_time': values_set['first_time'],
                        'first_time_name': values_set['first_time_name'],
                        'second_time': values_set['second_time'],
                        'second_time_name': values_set['second_time_name'],
                        'third_time_name': values_set['third_time_name'],
                        'use_auto': values_set['use_auto'],
                        'word_ticket': values_set['word_ticket'],
                        'word_order': values_set['word_order'],
                        'all_attributes': values_set['all_attributes']
                    }
                    with open('settings.ini', 'w', encoding='cp1251') as settings_file:
                        json.dump(setting_set, settings_file, ensure_ascii=False)
                    window_setting.close()
                    break

                elem_default_name.update(disabled=values_set['use_auto'])
                for i in elem_list:
                    i.update(disabled=not values_set['use_auto'])

        elif event == "Настройка остатков":
            layout_ostatki = [[sg.Text("Коэффициенты распределения: ")],
                              [sg.Text("Wildberries: "), sg.Input(key='wb_koef',
                                                                  default_text=float(ostatki_set['wb_koef']), size=3)],
                              [sg.Text("OZON: "), sg.Input(key='ozon_koef',
                                                                  default_text=ostatki_set['ozon_koef'], size=3)],
                              [sg.Text("Яндекс.Маркет: "), sg.Input(key='ya_koef',
                                                                  default_text=ostatki_set['ya_koef'], size=3)],
                              [sg.T("")], [sg.Text("Настройки файла остатков:")],
                              [sg.Text("Колонка с артикулами: "), sg.Input(key='art_kol',
                                                                  default_text=ostatki_set['art_kol'], size=3),
                              sg.Text("     Колонка с наименованием: "), sg.Input(key='nai_kol',
                                                                                default_text=ostatki_set['nai_kol'],
                                                                                size=3)],
                              [sg.Text("Колонка со штрихкодами: "), sg.Input(key='bar_kol',
                                                                                  default_text=ostatki_set['bar_kol'],
                                                                                  size=3),
                              sg.Text("     Колонка с количеством: "), sg.Input(key='kol_kol',
                                                                                    default_text=ostatki_set['kol_kol'],
                                                                                    size=3)],
                              [sg.Text("Неучитываемый минимум: "), sg.Input(key='predel',
                                                                                    default_text=ostatki_set['predel'],
                                                                                    size=3),
                              sg.Text("     Строк пропустить: "), sg.Input(key='skip_row',
                                                                                    default_text=ostatki_set['skip_row'],
                                                                                    size=3)],
                              [sg.Text("Название склада OZON:      "), sg.Input(key='name_sklad_ozon',
                                                                          default_text=ostatki_set['name_sklad_ozon'],
                                                                          size=45)],
                              [sg.Text("WB-артикул-баркод-фасовка"),
                               sg.Input(key="wb_art_bar_fas", default_text=ostatki_set['wb_art_bar_fas']),
                               sg.FileBrowse(button_text='Выбрать')],
                              [sg.Text("OZON-артикул-фасовка        "),
                               sg.Input(key="ozon_art_fas", default_text=ostatki_set['ozon_art_fas']),
                               sg.FileBrowse( button_text='Выбрать')],
                              [sg.Text("Я.Маркет-артикул-фасовка   "),
                               sg.Input(key="ya_art_fas", default_text=ostatki_set['ya_art_fas']),
                               sg.FileBrowse( button_text='Выбрать')],
                              [sg.Text("СБЕР-артикул-артикул-фасовка"),
                               sg.Input(key="sb_art_fas", default_text=ostatki_set['sb_art_fas']),
                               sg.FileBrowse(button_text='Выбрать')],
                              [sg.Text("Шаблон OZON:                    "),
                               sg.Input(key="shab_ozon", default_text=ostatki_set['shab_ozon']),
                               sg.FileBrowse(button_text='Выбрать')],
                              [sg.Text("Шаблон Я.Маркет:               "),
                               sg.Input(key="shab_ya", default_text=ostatki_set['shab_ya']),
                               sg.FileBrowse(button_text='Выбрать')],
                              [sg.Text("Шаблон СБЕР:                    "),
                               sg.Input(key="shab_sber", default_text=ostatki_set['shab_sber']),
                               sg.FileBrowse(button_text='Выбрать')],
                              [sg.Text("Файл подмены артикулов:   "),
                               sg.Input(key="artikul_swap", default_text=ostatki_set['artikul_swap']),
                               sg.FileBrowse(button_text='Выбрать')],
                              [sg.T("")], [sg.Button("Сохранить")]
                              ]


            window_setting = sg.Window('Настройки остатков', layout_ostatki, resizable=True, size=(620, 650))
            while True:
                event_ost, values_ost = window_setting.read()
                if event_ost == sg.WIN_CLOSED or event_ost == "Exit":
                    break
                elif event_ost == 'Сохранить':
                    ostatki_set = {
                        "wb_koef": values_ost["wb_koef"],
                        "ozon_koef": values_ost["ozon_koef"],
                        "ya_koef": values_ost["ya_koef"],
                        "art_kol": values_ost["art_kol"],
                        "nai_kol": values_ost["nai_kol"],
                        "kol_kol": values_ost["kol_kol"],
                        "skip_row": values_ost["skip_row"],
                        "bar_kol": values_ost["bar_kol"],
                        "predel": values_ost["predel"],
                        "name_sklad_ozon": values_ost["name_sklad_ozon"],
                        "wb_art_bar_fas": values_ost["wb_art_bar_fas"],
                        "ozon_art_fas": values_ost["ozon_art_fas"],
                        "ya_art_fas": values_ost["ya_art_fas"],
                        "sb_art_fas": values_ost["sb_art_fas"],
                        "shab_ozon": values_ost["shab_ozon"],
                        "shab_ya": values_ost["shab_ya"],
                        "shab_sber": values_ost["shab_sber"],
                        "artikul_swap": values_ost["artikul_swap"]
                    }
                    with open('settings_ost.ini', 'w', encoding='cp1251') as settings_file:
                        json.dump(ostatki_set, settings_file, ensure_ascii=False)
                    window_setting.close()
                    break

        elif event == "Настройка проверки остатков":
            layout_prod_ost = [[sg.Text("Пропустить строк: "), sg.Input(key='prod_skip_row',
                                                                        default_text=prod_set['prod_skip_row'],
                                                                        size=8)],
                               [sg.Text("Даты находятся в колонке: "), sg.Input(key='data_column',
                                                                                default_text=prod_set['data_column'],
                                                                                size=8)],
                               [sg.Text("За сколько дней продажи: "), sg.Input(key='data_',
                                                                               default_text=prod_set['data_'],
                                                                               size=8)],
                               [sg.Text("Колонка с наименованием"), sg.Input(key='nai', default_text=prod_set['nai'],
                                                                               size=8),
                               sg.Text("Колонка с артикулом"), sg.Input(key='art', default_text=prod_set['art'],
                                                                            size=8)],
                               [sg.Text("Колонка с баркодом"), sg.Input(key='bar', default_text=prod_set['bar'],
                                                                            size=8),
                               sg.Text("Колонка с продажами"), sg.Input(key='prod', default_text=prod_set['prod'],
                                                                               size=8)],
                               [sg.Text("Колонки по которым группируем продажи"),
                                sg.Input(key='priority_column', default_text=prod_set['priority_column'],
                                        size=8)],
                               [sg.Text("Специальный разделитель для указания разных фасовок"),
                               sg.Input(key='delim', default_text=prod_set['delim'], size=8)],
                               [sg.Checkbox("Включить указание фасовки в наименовании. Формат: (xxxx)",
                                            key="fas",  default=prod_set['fas'])],
                               [sg.Text("Путь выходного файла:")],
                               [sg.Text(prod_set['prod_out_path']),
                                sg.FolderBrowse(key='prod_out_path', initial_folder=prod_set['prod_out_path'], )],
                               [sg.Button("Сохранить")]
            ]

            window_setting = sg.Window('Настройка проверки остатков', layout_prod_ost, resizable=True, size=(620, 400))
            while True:
                event_set_prod, values_set_prod = window_setting.read()
                if event_set_prod == sg.WIN_CLOSED or event_set_prod == "Exit":
                    break
                elif event_set_prod == 'Сохранить':
                    prod_set = {
                        "prod_skip_row": values_set_prod["prod_skip_row"],
                        "data_column": values_set_prod["data_column"],
                        "data_": values_set_prod["data_"],
                        "nai": values_set_prod["nai"],
                        "art": values_set_prod["art"],
                        "bar": values_set_prod["bar"],
                        "prod": values_set_prod["prod"],
                        "priority_column": values_set_prod["priority_column"],
                        "delim": values_set_prod["delim"],
                        "fas": values_set_prod["fas"],
                        "prod_out_path": values_set_prod["prod_out_path"] if values_set_prod["prod_out_path"] else prod_set['prod_out_path']
                    }
                    with open('settings_ost_prod.ini', 'w', encoding='cp1251') as settings_file:
                        json.dump(prod_set, settings_file, ensure_ascii=False)
                    window_setting.close()
                    break

        elif event == "Анализ заказов":
            brand_list = []
            art_list = []
            check_brand_list = []
            check_art_list = []
            layout_analiz = [[sg.Text('Файл с заказами:'), sg.Input(),
                              sg.FileBrowse(key='zakaz', button_text='Выбрать', file_types=(('Excell', '*.xlsx'),))],
                             [sg.Button('Выбрать бренд'), sg.Button('Выбрать артикул'), sg.Text(key="check_text")],
                             [sg.Text(f'Выбранные бренды:', key="text_brand")],
                             [sg.Text(f'Выбранные артикулы:', key="text_art")],
                             [sg.Radio('Артикул', group_id='art_or_brand', key='Артикул продавца', default=True),
                              sg.Radio('Бренд', group_id='art_or_brand', key='Бренд', default=False)],
                             [sg.Radio('Сумма заказов минус комиссия WB, руб.', group_id='col_or_sum',
                                 key='Сумма заказов минус комиссия WB, руб.', default=True),
                              sg.Radio('Заказано, шт.', group_id='col_or_sum', key='Заказано, шт.',  default=False)],
                             [sg.Text("От:"), sg.Input(size=(12,1), key='begin_date'),
                              sg.CalendarButton('дата', format='%d.%m.%Y')],
                             [sg.Text("До:"), sg.Input(size=(12, 1),key='end_date'),
                              sg.CalendarButton('дата', format='%d.%m.%Y')],
                             [sg.Button('Сформировать график'), sg.Checkbox('Суммировать заказы', key='sum_order',
                                                                            default=False)],
                             [sg.T()],[sg.T()],
                             [sg.T('Введите лимит'), sg.Input('0', key='-LIMIT-', size=12),
                              sg.Checkbox('Меньше', default=False, key='-BIGGER_SMALLER-'),
                              sg.Combo(['Рубли', 'Штуки'], key='-RUBLI-SHTUKI-', default_value='Рубли'),
                              sg.Button('Проверить заказы')],

            ]

            window_analiz = sg.Window('Анализ заказов', layout_analiz, resizable=True, size=(620, 600))
            while True:
                event_analiz, values_analiz = window_analiz.read()

                if event_analiz == 'Выбрать бренд':
                    try:
                        df, min_day, max_day = analitics.open_database(values_analiz['zakaz'])
                        brand_list = analitics.get_unique_values(df, 'Бренд')
                        art_list = analitics.get_unique_values(df, 'Артикул продавца')

                    except Exception as P:
                        sg.Print(P)
                    layout_brand = [[sg.Button('Выбрать всё'), sg.Button('Очистить выбор'), sg.Button('Ok')],
                        [sg.Column(layout=[[sg.Checkbox(i, key=str(i), default=(True if i in check_brand_list else False),
                                      enable_events=True)] for i in brand_list], size=(800, 850),
                                            scrollable=True,  vertical_scroll_only=True)]]
                    window_brand = sg.Window('Список брендов', layout_brand, resizable=True, size=(800, 850))
                    while True:
                        event_brand, values_brand = window_brand.read()
                        if event_brand == sg.WIN_CLOSED:
                            break
                        elif event_brand == 'Ok':
                            check_brand_list = []
                            elem_br = window_analiz['text_brand']
                            for key, values in values_brand.items():
                                if values:
                                    check_brand_list.append(key)
                            elem_br.update(value=str(check_brand_list))
                            window_brand.close()
                        elif event_brand == 'Очистить выбор' or event_brand == 'Выбрать всё':
                            for i in brand_list:
                                el = window_brand[i]
                                if event_brand == 'Очистить выбор':
                                    el.update(value=False)
                                else:
                                    el.update(value=True)

                elif event_analiz == 'Выбрать артикул':
                    try:
                        df, min_day, max_day = analitics.open_database(values_analiz['zakaz'])
                        brand_list = analitics.get_unique_values(df, 'Бренд')
                        art_list = analitics.get_unique_values(df, 'Артикул продавца', brand=check_brand_list)
                    except Exception as P:
                        sg.Print(P)
                    if len(settings['all_attributes']) > 0:
                        art_name_dict = analitics.connected_art_with_name(settings['all_attributes'])
                        art_name_list = [[i, *art_name_dict[i]] if i in art_name_dict else
                                         [i, *[' ' for p in range(4)]] for i in art_list]
                        art_name_list.sort(key=lambda x: (x[1], x[2]))
                        layout_art = [[sg.Button('Выбрать всё'), sg.Button('Очистить выбор'), sg.Button('Ok')],
                            [sg.Column([[sg.Checkbox(' '.join(art_name_list[i]), key=str(art_name_list[i][0]),
                                          default=(True if art_name_list[i][0] in check_art_list else False))]
                             for i in range(len(art_name_list))], size=(800, 850), scrollable=True,
                                       vertical_scroll_only=True)]
                        ]
                    else:
                        layout_art = [[sg.Button('Выбрать всё'), sg.Button('Очистить выбор'), sg.Button('Ok')],
                            [sg.Column(layout=[[sg.Checkbox(
                                i, key=str(i), default=(True if i in check_art_list else False))] for i in art_list],
                                size=(800, 850), scrollable=True,  vertical_scroll_only=True)]
                        ]
                    window_art = sg.Window('Список артикулов', layout_art, resizable=True, size=(800, 850))

                    while True:
                        event_brand, values_brand = window_art.read()
                        if event_brand == sg.WIN_CLOSED:
                            break
                        elif event_brand == 'Ok':
                            check_art_list = []
                            elem_art = window_analiz['text_art']
                            for key, values in values_brand.items():
                                if values:
                                    check_art_list.append(key)
                            elem_art.update(value=str(check_art_list))
                            window_art.close()
                        elif event_brand == 'Очистить выбор' or event_brand == 'Выбрать всё':
                            for i in art_list:
                                el = window_art[i]
                                if event_brand == 'Очистить выбор':
                                    el.update(value=False)
                                else:
                                    el.update(value=True)

                elif event_analiz == 'Сформировать график':
                    try:
                        need_columns = []
                        n_c = ['Артикул продавца', 'Бренд', 'Сумма заказов минус комиссия WB, руб.', 'Заказано, шт.']
                        for i in n_c:
                            if values_analiz[i]:
                                need_columns.append(i)
                        if len(values_analiz['begin_date']) > 0:
                            min_day = values_analiz['begin_date']
                            min_day = datetime.datetime.strptime(min_day, '%d.%m.%Y')
                        if len(values_analiz['end_date']) > 0:
                            max_day = values_analiz['end_date']
                            max_day = datetime.datetime.strptime(max_day, '%d.%m.%Y')
                        period = [min_day, max_day]
                        graph = []
                        days_axis = []
                        if 'Артикул продавца' in need_columns:
                            item = check_art_list
                        elif 'Бренд' in need_columns:
                            item = check_brand_list
                        if len(item) < 1:
                            sg.Print('Выбирете хотя бы 1 атрибут')
                            continue
                        for i in item:
                            df_current, day_axis = analitics.make_axis(df, period, need_columns, i)
                            graph.append(df_current)
                            days_axis.append(day_axis)
                            min_day = False
                            max_day = False
                        for i in days_axis:
                            current_min = min(i)
                            current_max = max(i)
                            if not min_day:
                                min_day = current_min
                            if not max_day:
                                max_day = current_max
                            if current_min < min_day:
                                min_day = current_min
                            if current_max > max_day:
                                max_day = current_max
                        days_axis = pd.date_range(min_day, max_day)
                        if len(settings['all_attributes']) > 0 and 'Бренд' not in need_columns:
                            art_name_dict = analitics.connected_art_with_name(settings['all_attributes'])
                            analitics.make_graph(graph, days_axis, art_name=art_name_dict,
                                                 sum_flag=values_analiz['sum_order'])
                        else:
                            analitics.make_graph(graph, days_axis, sum_flag=values_analiz['sum_order'])
                    except Exception as P:
                        sg.Print(P)

                elif event_analiz == 'Проверить заказы':
                    try:
                        if len(values_analiz['begin_date']) > 0:
                            min_day = values_analiz['begin_date']
                            min_day = datetime.datetime.strptime(min_day, '%d.%m.%Y')
                        if len(values_analiz['end_date']) > 0:
                            max_day = values_analiz['end_date']
                            max_day = datetime.datetime.strptime(max_day, '%d.%m.%Y')
                        period = [min_day, max_day]
                        rubli_shtuki_flag = values_analiz['-RUBLI-SHTUKI-']
                        limit = int(values_analiz['-LIMIT-'])
                        art_name_dict = analitics.connected_art_with_name(settings['all_attributes'])
                        analitics.check_main_file_on_limits(values_analiz['zakaz'], limit,
                                                            values_analiz['-BIGGER_SMALLER-'],
                                                            art_name_dict, period, rubli_shtuki_flag,
                                                            settings['all_attributes'])
                        sg.Print('Завершено. Файл сохранён в корневую директорию')
                    except Exception as P:
                        sg.Print(P)


                elif event_analiz == sg.WIN_CLOSED or event_analiz == "Exit":
                    break

        elif event == 'Работа с заказами':
            event_orders_list = []
            current_select = 0
            fixed_time = 'Заказы не зафиксированы'
            def make_window(file):
                current_select = 0
                orders_list = api.get_orders(file, fixed_time)
                supplies_list = api.get_supplies()
                len_sullies_dict = {}
                for i in supplies_list:
                    len_sullies_dict[i[0]] = str(len(api.get_info_supply(i[0]).json()['orders']))
                sg.PrintClose()

                layout_orders = [[sg.Button("Добавить задания к поставке"), sg.Button("Выделить всё"), sg.Input(key='make_supply'),
                                  sg.Button("Создать поставку"), sg.T(f'{current_select}', key='current_select')],
                                 [sg.Button('Зафиксировать' if fixed_time == 'Заказы не зафиксированы' else 'Снять фиксацию'),
                                  sg.T(fixed_time),
                                  sg.T('                                                                                   '),
                                  sg.Button('Получить этикетки')],
                                 [sg.Column(layout=[[sg.Checkbox(' '.join(i[1]), key=str("order_" + str(i[0])), default=False,
                                          enable_events=True)] for i in orders_list], size=(600, 850),
                                            scrollable=True,  vertical_scroll_only=True),
                                  sg.Column(layout=[[sg.Radio(i[1] + ' // ' + len_sullies_dict[i[0]], group_id='radio_supply', key=str("supply_" + str(i[0])))]
                                                    for i in supplies_list], size=(200, 850),
                                            scrollable=True, vertical_scroll_only=True)
                                  ]
                                 ]
                window_orders = sg.Window('Работа с заказами', layout_orders, resizable=True, size=(835, 900),
                                          return_keyboard_events=True)
                orders_list = [i[0] for i in orders_list]
                return window_orders, orders_list

            try:
                all_atributes = settings['all_attributes']
                window_orders, orders_list = make_window(all_atributes)
            except Exception as P:
                sg.Print(P)
                sg.Print('не удалось получить списки заказов и поставок')
                time.sleep(7)
                continue
            while True:
                event_orders, values_orders = window_orders.read()
                if event_orders == sg.WIN_CLOSED:
                    break
                elif event_orders == 'Получить этикетки':
                    for key, value in values_orders.items():
                        if "supply_" in key and value == True:
                            supply_id = key[7:]
                            supply_name = window_orders[key].Text.split(' // ')

                            skus_name_dict = analitics.connected_skus_with_name(all_atributes)

                            supply_list = api.get_orders_into_supply(supply_id)
                            for ind, i in enumerate(supply_list):
                                if i[3] in skus_name_dict.keys():
                                    supply_list[ind].append(' '.join(skus_name_dict[i[3]]))
                                else:
                                    for key in skus_name_dict.keys():
                                        if i[3] in key:
                                            supply_list[ind].append(' '.join(skus_name_dict[key]))
                            without_name_items_list = []
                            for ind in range(len(supply_list)):
                                if len(supply_list[ind]) < 5:
                                    without_name_items_list.append(supply_list[ind])
                            supply_list = [i for i in supply_list if i not in without_name_items_list]
                            supply_list.sort(key=lambda x: x[4])
                            supply_list += without_name_items_list

                            c = canvas.Canvas(f'{supply_name[0]}.pdf', pagesize=(400, 400))

                            for i in supply_list:
                                image_ = api.get_labels_of_order(i)
                                if len(i) == 5:
                                    label_image = image_change.image_show(image_, name=i[4], art=i[2])
                                else:
                                    label_image = image_change.image_show(image_, name=None, art=i[2])
                                image_pil = Image.fromarray(label_image)
                                image = ImageReader(image_pil)

                                c.drawImage(image, 0, 0)
                                c.showPage()
                                if len(i) == 5:
                                    sg.Print(i[4])
                                else:
                                    sg.Print(i[2])

                            c.save()

                            sg.Print('Файл создан')


                elif event_orders == 'Зафиксировать':
                    id_name_list = api.get_orders(all_atributes, fixed_time)
                    id_name_list.sort(key=lambda x: (x[2]))
                    fixed_time = datetime.datetime.fromisoformat(id_name_list[-1][2])
                    window_orders.close()
                    window_orders, orders_list = make_window(all_atributes)
                elif event_orders == 'Снять фиксацию':
                    fixed_time = 'Заказы не зафиксированы'
                    window_orders.close()
                    window_orders, orders_list = make_window(all_atributes)

                elif event_orders == 'Выделить всё':
                    if window_orders['Выделить всё'].ButtonText == 'Выделить всё':
                        current_select = 0
                        for key in values_orders.keys():
                            if "order_" in key:
                                window_orders[key].update(value=True)
                                window_orders['Выделить всё'].update(text='Снять выделение')
                                current_select += 1
                        window_orders['current_select'].update(value=current_select)
                    elif window_orders['Выделить всё'].ButtonText == 'Снять выделение':
                        for key in values_orders.keys():
                            if "order_" in key:
                                window_orders[key].update(value=False)
                                window_orders['Выделить всё'].update(text='Выделить всё')
                        window_orders['current_select'].update(value=0)
                elif event_orders == 'Создать поставку':
                    if len(values_orders['make_supply']) == 0:
                        sg.Print('Введите имя поставки')
                        continue
                    try:
                        api.make_supply(values_orders['make_supply'])
                    except Exception as P:
                        sg.Print(P)
                        sg.Print("Не удалось создать поставку")
                        time.sleep(7)
                        break
                    time.sleep(5)
                    window_orders.close()
                    window_orders, orders_list = make_window(all_atributes)
                elif event_orders == 'Добавить задания к поставке':
                    orders = []
                    supply_id = ''
                    for key, value in values_orders.items():
                        if "order_" in key and value == True:
                            orders.append(key[6:])
                        if "supply_" in key and value == True:
                            supply_id = key[7:]
                    if len(supply_id) == 0:
                        sg.Print('Выберите поставку')
                        continue
                    for i in orders:
                        try:
                            api.add_order_into_supply(i, supply_id)
                        except Exception as P:
                            sg.Print(P)
                            sg.Print("Не удалось добавить задание к поставке")
                            time.sleep(7)
                            break
                    time.sleep(5)
                    window_orders.close()
                    window_orders, orders_list = make_window(all_atributes)
                elif event_orders:
                    current_select = 0
                    for key, value in values_orders.items():
                        if "order_" in key and value == True:
                            current_select += 1
                    window_orders['current_select'].update(value=current_select)
                    if len(event_orders_list) == 0:
                        pass
                    elif len(event_orders_list) > 0 and event_orders_list[-1] == 'Shift_L:16':
                        flag = False
                        if 'order_' in event_orders:
                            for i in event_orders_list[::-1]:
                                if 'order_' in i:
                                    last_event_order = i[6:]
                                    last_index = orders_list.index(int(last_event_order))
                                    flag = True
                                    if 'order_' in event_orders:
                                        current_event_orders = event_orders[6:]
                                        current_index = orders_list.index(int(current_event_orders))
                                    else:
                                        break
                                    if flag and current_index > last_index:
                                        diapazon = (last_index+1, current_index)
                                    else:
                                        diapazon = (last_index-1, current_index, -1)

                                    for i in range(*diapazon):
                                        if values_orders['order_' + str(orders_list[i])]:
                                            current_select -= 1
                                        else:
                                            current_select += 1
                                        window_orders['order_' + str(orders_list[i])].update(
                                            value=(False if values_orders['order_' + str(orders_list[i])] == True else True))
                                    window_orders['current_select'].update(value=current_select)
                                    break
                    event_orders_list.append(event_orders)

        elif event == 'Изменить цены':
            price_layout = [
                [sg.Text("Файл с товарами"), sg.Input(), sg.FileBrowse(key='tovar', button_text='Выбрать',
                                                         file_types=(('xlsl', ('*.xlsx')),))],
                [sg.Text("Файл с условиями"), sg.Input(), sg.FileBrowse(key='uslovia', button_text='Выбрать',
                                                         file_types=(('xlsl', ('*.xlsx')),))],
                [sg.Combo(['WB', 'Ozon', 'Market', 'Sber', 'All'], key='flag'),
                 sg.Checkbox('Изменить цены по API', key='-API-', default=False)],
                [sg.Button('Изменить цены')]
            ]
            price_window = sg.Window('Установка цены', price_layout, resizable=True, size=(680, 680))

            while True:
                event_price, values_price = price_window.read()
                if event_price == sg.WIN_CLOSED:
                    break
                elif event_price == 'Изменить цены':
                    # try:
                        tovar, uslovia = values_price['tovar'], values_price['uslovia']
                        marketplace = values_price['flag']
                        api_enabled = values_price['-API-']
                        if marketplace == 'WB':
                            token_wb = open('price_token.txt').read()
                            change_price.get_wb_price(tovar, uslovia, api_enabled, token_wb)
                        elif marketplace == 'Ozon':
                            token_ozon = open('price_token_ozon.txt').read()
                            change_price.get_ozon_price(tovar, uslovia, api_enabled, token_ozon)
                        elif marketplace == 'Market':
                            token_dict = json.load(open('token_yandex.txt'))
                            token_ya = token_dict['Token']
                            bussines_id = token_dict['Bussines_id']
                            change_price.get_yandex_price(tovar, uslovia, api_enabled, token_ya, bussines_id)
                        elif marketplace == 'Sber':
                            change_price.get_sber_price(tovar, uslovia, api_enabled)

                        elif marketplace == 'All':
                            token_wb = open('price_token.txt').read()
                            token_ozon = open('price_token_ozon.txt').read()
                            token_dict = json.load(open('token_yandex.txt'))
                            token_ya = token_dict['Token']
                            bussines_id = token_dict['Bussines_id']
                            change_price.get_wb_price(tovar, uslovia, api_enabled, token_wb)
                            change_price.get_ozon_price(tovar, uslovia, api_enabled, token_ozon)
                            change_price.get_yandex_price(tovar, uslovia, api_enabled, token_ya, bussines_id)
                            change_price.get_sber_price(tovar, uslovia, api_enabled)
                        sg.Print('Изменение цен завершено. Файлы для копирования сохранены в папку с приложением')
                    # except Exception as P:
                    #     sg.Print(P)

        elif event == 'Изм. габариты WB':
            layout_change_sizes_wb = [[
                sg.T('Отчёт по коэф. логистики'), sg.Input(),
                sg.FilesBrowse(key="-REPORT_FILE-", button_text='Выбрать', file_types=(('XLSX', ('*xlsx')),))],
                [sg.T('Файл с товарами'), sg.Input(),
                 sg.FilesBrowse(key="-TOVAR-", button_text='Выбрать', file_types=(('XLSX', ('*xlsx')),))],
                [sg.T(), sg.Checkbox('Внести изменения в файл с товарами', key='-CHANGE_FILE_TOVAR-', default=True)],
                [sg.Button(button_text='Изменить габариты', key='-START-'),
                 sg.Button(button_text='Изменить габариты на основании основного фалйа', key='-START2-')]
            ]
            window_change_sizes_wb = sg.Window('Изменить габариты WB', layout_change_sizes_wb, resizable=True,
                                               size=(650, 250))
            while True:
                event_change_sizes_wb, values_change_sizes_wb = window_change_sizes_wb.read()
                if event_change_sizes_wb == sg.WIN_CLOSED or event == "Exit":
                    break
                elif event_change_sizes_wb == '-START-':
                    content_token = open('content_token.txt').read()
                    report = pd.read_excel(values_change_sizes_wb['-REPORT_FILE-'])
                    art_nmID_set = change_size_wb.art_nmID(settings['all_attributes'])

                    sg.Print('Новые габариты')
                    new_cards_list = [[],[],[],[]]
                    new_cards_ind = 0
                    cards, updatedAt, nm_ID = change_size_wb.get_card(content_token)
                    if not cards:
                        sg.Print('Не удалось пролучить список товаров')
                        time.sleep(4)
                        break
                    else:
                        new_cards = cards
                        cards = cards['cards']
                        while len(new_cards['cards']) != 0:
                            new_cards, updatedAt, nm_ID = change_size_wb.get_card(content_token, updatedAt, nm_ID)
                            cards += new_cards['cards']

                    if values_change_sizes_wb['-CHANGE_FILE_TOVAR-']:
                        tovar_df = pd.read_excel(values_change_sizes_wb['-TOVAR-'])
                    for ind, row in report.iterrows():
                        nmID = row['Номенклатура']
                        sizes = change_size_wb.get_num_sizes(report, nmID)
                        if values_change_sizes_wb['-CHANGE_FILE_TOVAR-']:

                            shirina_etal = tovar_df['Ширина упаковки эталон'][tovar_df['Артикул WB'] == nmID]
                            dlina_etal = tovar_df['Длина упаковки эталон'][tovar_df['Артикул WB'] == nmID]
                            vysota_etal = tovar_df['Высота упаковки эталон'][tovar_df['Артикул WB'] == nmID]

                            ob_real = sizes[0] * sizes[1] * sizes[2]
                            ob_etal = shirina_etal.values[0] * dlina_etal.values[0] * vysota_etal.values[0]
                            if ob_real > ob_etal:
                                tovar_df['Ширина упаковки WB'][tovar_df['Артикул WB'] == nmID] = sizes[0]
                                tovar_df['Длина упаковки WB'][tovar_df['Артикул WB'] == nmID] = sizes[1]
                                tovar_df['Высота упаковки WB'][tovar_df['Артикул WB'] == nmID] = sizes[2]

                            else:
                                tovar_df['Ширина упаковки WB'][tovar_df['Артикул WB'] == nmID] = shirina_etal
                                sizes[0] = int(shirina_etal.values[0])
                                tovar_df['Длина упаковки WB'][tovar_df['Артикул WB'] == nmID] = dlina_etal
                                sizes[1] = int(dlina_etal.values[0])
                                tovar_df['Высота упаковки WB'][tovar_df['Артикул WB'] == nmID] = vysota_etal
                                sizes[2] = int(vysota_etal.values[0])

                        if nmID in art_nmID_set.keys():
                            art = art_nmID_set[nmID]
                        else:
                            sg.Print(
                                f'{nmID} - Не найден в файле общих характеристик. Скачайте свежую версию файла и попробуйте снова \n'
                                f'Габариты {nmID} не обновлены')
                            continue

                        new_sizes = change_size_wb.change_size(art, cards, sizes)
                        if new_sizes:
                            if len(new_cards_list[new_cards_ind]) == 100:
                                new_cards_ind += 1
                            new_cards_list[new_cards_ind].append(new_sizes)
                            sg.Print(f'{art} - {sizes}')
                        else:
                            sg.Print(f'Артикул {art} не найден в полученных карточках товара' )
                            break
                    for i in new_cards_list:
                        if len(i) > 0:
                            post_sizes = change_size_wb.post_size(content_token, i)
                            if post_sizes.status_code == 200:
                                sg.Print('Габариты на сайте успешно обновлены')

                    sg.Print('Завершено')
                    if values_change_sizes_wb['-CHANGE_FILE_TOVAR-']:
                        tovar_df.to_excel(values_change_sizes_wb['-TOVAR-'], index=False)
                        sg.Print('Файл с товарами перезаписан')

                elif event_change_sizes_wb == '-START2-':
                    content_token = open('content_token.txt').read()
                    cards, updatedAt, nm_ID = change_size_wb.get_card(content_token)
                    if not cards:
                        sg.Print('Не удалось пролучить список товаров')
                        time.sleep(4)
                        break
                    else:
                        new_cards = cards
                        cards = cards['cards']
                        while len(new_cards['cards']) != 0:
                            new_cards, updatedAt, nm_ID = change_size_wb.get_card(content_token, updatedAt, nm_ID)
                            cards += new_cards['cards']

                    tovar_df = pd.read_excel(values_change_sizes_wb['-TOVAR-'], converters={'Артикул WB': str})
                    new_df = tovar_df[['Артикул продавца WB', 'Ширина упаковки WB', 'Длина упаковки WB', 'Высота упаковки WB']].dropna()
                    arts = new_df['Артикул продавца WB'].unique()
                    new_cards_list = [[], [], [], []]
                    new_cards_ind = 0

                    for i in arts:
                        s = int(new_df['Ширина упаковки WB'][new_df['Артикул продавца WB'] == i].values[0])
                        d = int(new_df['Длина упаковки WB'][new_df['Артикул продавца WB'] == i].values[0])
                        v = int(new_df['Высота упаковки WB'][new_df['Артикул продавца WB'] == i].values[0])
                        sizes = [s, d, v]

                        new_sizes = change_size_wb.change_size(i, cards, sizes)
                        if new_sizes:
                            if len(new_cards_list[new_cards_ind]) == 100:
                                new_cards_ind += 1
                            new_cards_list[new_cards_ind].append(new_sizes)
                            sg.Print(f'{i} - {sizes}')
                        else:
                            sg.Print(f'Артикул {i} не найден в полученных карточках товара' )
                    for i in new_cards_list:
                        if len(i) > 0:
                            post_sizes = change_size_wb.post_size(content_token, i)
                            if post_sizes.status_code == 200:
                                sg.Print('Габариты на сайте успешно обновлены')



        try:
            not_ozon = not values['ozon']
            elem.update(disabled=not_ozon)
            last_order.update(disabled=not_ozon)
        except:
            pass
        settings = open('settings.ini', 'r')
        settings = json.load(settings)

        try:
            compile_text = check_time(settings, settings['use_datetime'])
        except Exception as P:
            compile_text = 'Pass'
            sg.Print(P)
            sg.Print('Нарушен формат времени. Советуем использовать ЧЧ.ММ.СС\n'
                     'Не ставьте пробелы. Если отчность до секунд или минут не нужна - указывайте только часы или часы с минутами')

        default_text = f'{compile_text}' if settings["use_auto"] == True else settings["default_name"]
        name_elem.update(value=default_text)

        # else:
        #     sg.Print(response.text)
        #     time.sleep(20)
    # except Exception as P:
    #     sg.Print('Не удалось подключиться к серверу')
    #     sg.Print(P)
    #     time.sleep(10)