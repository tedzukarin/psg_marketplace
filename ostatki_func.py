import time

import pandas as pd
import PySimpleGUI as sg
import json
import os.path
import openpyxl
from copy import copy
import api


def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


def raspr_ost(ostat, rezerv=None, sber_flag=False, use_api_flag=False):
    with open('settings_ost.ini', 'r') as setting_file:
        setting = json.load(setting_file)

    path, file_name = os.path.split(ostat)

    wb_koef = setting['wb_koef']
    ozon_koef = setting['ozon_koef']
    ya_koef = setting['ya_koef']
    art_kol = setting['art_kol']
    nai_kol = setting['nai_kol']
    kol_kol = setting['kol_kol']
    skip_row = setting['skip_row']
    bar_kol = setting['bar_kol']
    predel = setting['predel']
    name_sklad_ozon = setting['name_sklad_ozon']
    wb_art_bar_fas = setting['wb_art_bar_fas']
    ozon_art_fas = setting['ozon_art_fas']
    ya_art_fas = setting['ya_art_fas']
    sb_art_fas = setting['sb_art_fas']
    shab_ozon = setting['shab_ozon']
    shab_ya = setting['shab_ya']
    shab_sber = setting['shab_sber']
    artikul_swap = setting['artikul_swap']


    if artikul_swap:
        with open(artikul_swap, 'r', encoding='cp1251') as artikul_swap:
            swap = json.load(artikul_swap)
            if 'WB' in swap.keys():
                wb_swap = swap['WB']
            if 'OZON' in swap.keys():
                ozon_swap = swap['OZON']
            if 'YA' in swap.keys():
                ya_swap = swap['YA']

    if predel:
        predel = int(predel)
    else:
        predel = 0

    total = []
    ozon_art = []
    ya_art = []
    sb_art = []

    ostatki_xlsx = pd.read_excel(ostat)
    if rezerv:
        rezerv_xlsx = pd.read_excel(rezerv, skiprows=7, converters={'Количество': int})

    ostatki = []
    n = 0
    if skip_row:
        skip_row = int(skip_row)
    else:
        skip_row = 0
    for i, row in ostatki_xlsx.iterrows():
        n += 1
        if n > skip_row:
            art = row[int(art_kol)-1]
            kol = row[int(kol_kol)-1]
            name = row[int(nai_kol)-1]
            if rezerv and art in rezerv_xlsx['Код'].unique():
                kol = kol - rezerv_xlsx['Количество'][rezerv_xlsx['Код'] == art].values[0]
            ostatki.append([art, kol, name])

    nulls_and_minus = [['Наименование', 'Артикул', 'Баркод', 'фактический остаток']]
    wb_ostatki = [['Баркод', 'Количество']]
    ozon_ostatki = [['Склад', 'Артикул', 'Имя', 'Остатки']]
    ya_ostatki = [['Ошибки', 'Предупреждения', 'Ваш SKU *', 'Название товара', 'Доступное количество товара *'],
                  ['Заполняется автоматически. Здесь будет перечислен список ошибок для каждого товара после того, как вы загрузите шаблон в систему',
                   'Заполняется автоматически. Здесь будет перечислен список предупреждений для каждого товара после того, как вы загрузите шаблон в систему',
                   'Уникальный идентификатор товара. Обязательное поле.',
                   'По схеме: тип товара + бренд или производитель + модель + отличительные характеристики.',
                   'Общее количество товара, доступное для продажи на маркетплейсе и зарезервированное под заказы. Обязательное поле.']]
    sb_ostatki = [['id', 'Доступность товара', 'Категория', 'Производитель (Бренд)', 'Артикул', 'Модель', 'Название',
                   'Цена(руб)', 'Старая цена(руб)', 'Остаток', 'НДС', 'Штрихкод', 'Ссылка на картинку',' Описание',
                   'Ссылка на товар на сайте магазина', 'Время заказа До', 'Дней на отгрузку'],]

    if sb_art_fas and sber_flag:
        try:
            sb_art_file = open(sb_art_fas, encoding='cp1251').read().splitlines()
            sb_api = []
            for i in sb_art_file:
                sb_art.append(i.split('\t'))

            '''Получаем датафрейм из листа и меняем остатки'''
            df = pd.read_excel(shab_sber, skiprows=1, sheet_name='Список товаров')
            sb_api_list = []
            for index, row in df.iterrows():
                kol = 0
                fas = 1
                vendor_code = row['vendor_code']
                for j in sb_art:
                    if j[0] == vendor_code:
                        one_c_code, fas = j[1], int(j[2])
                        for c in ostatki:
                            if c[0] == one_c_code:
                                kol = int(c[1])
                                break
                art_ost = round(kol / fas)
                df['instock'][df['vendor_code'] == vendor_code] = art_ost
            columns = df.columns.tolist()
            table = df.values.tolist()
            sb_ostatki.append(columns)


            for i in table:
                sb_ostatki.append(i)

            new_df = pd.DataFrame(data=sb_ostatki[1:], columns=sb_ostatki[0])
            new_df.to_excel(f'{path}//sb.xlsx', sheet_name='Список товаров', index=False)

            '''Копируем страницу Инструкция'''
            wb_target = openpyxl.load_workbook(f'{path}//sb.xlsx')
            sheets = wb_target.sheetnames
            target_sheet = wb_target.create_sheet('Инструкция')
            wb_source = openpyxl.load_workbook(shab_sber, data_only=True)
            source_sheet = wb_source['Инструкция']
            copy_sheet(source_sheet, target_sheet)
            wb_target.move_sheet(sheets[0], offset=len(sheets))
            wb_target.save(f'{path}//sb.xlsx')
            wb_target.close()
            sg.Print('Остатки SBER записаны')

            if use_api_flag:
                for i in sb_art:
                    for c in ostatki:
                        if c[0] == i[1]:
                            fas = int(i[2])
                            kol = int(c[1])
                            art_ost = round(kol / fas)
                            sb_api.append([i[0], art_ost])
                try:
                    token = open('sber_token.txt').read()
                    response = api.send_ostatki_sb(sb_api, token)
                    if response.json()['success'] == 1:
                        sg.P('Остатки SBER обновлены')
                    else:
                        sg.P('Не удалось обновить остатки по API')
                except:
                    sg.P('Не удалось обновить остатки по API')

        except Exception as P:
            sg.Print(P)

    if wb_art_bar_fas and not sber_flag:
        total_file = open(wb_art_bar_fas, encoding='cp1251').read().splitlines()
        for i in total_file:
            p = i.split('\t')
            total.append([p[0], p[1], int(p[2])])

        for i in total:
            art, bar, fas = i[0], i[1], i[2]
            flag = False
            name_art = art
            if 'wb_swap' in locals() and art in wb_swap.keys():
                art = wb_swap[art]
            for p in ostatki:
                nai = p[2]
                if p[0] == art:
                    if type(p[1]) is int:
                        ost = p[1] // fas
                        ost = round(ost * float(wb_koef)) - predel
                    elif type(p[1]) is float:
                        sg.Print(f'WB-остатки. Проверьте артикул: {art} - дробное или нулевое значение остатка')
                        continue
                    if ost <= 0:
                        ost = 0
                        nulls_and_minus.append([nai, name_art, bar, ost])
                    wb_ostatki.append([bar, ost])
                    flag = True
                    break
            if not flag:
                wb_ostatki.append([bar, 0])
                nulls_and_minus.append([nai, name_art, bar, 0])
        try:
            wb_ostatki_to_df = pd.DataFrame(wb_ostatki[1:], columns=[wb_ostatki[0][0], wb_ostatki[0][1]])
            wb_ostatki_to_df.to_excel(f'{path}/wb.xlsx', index=False, sheet_name='Остатки')
            sg.Print('Остатки WB записаны')

        except Exception as ex:
            sg.Print(ex)

        if use_api_flag:
            try:
                api.send_ostatki_wb(wb_ostatki)
            except:
                sg.P('Не удалось обновить остатки по API')

    if ozon_art_fas and not sber_flag:
        ozon_art_file = open(ozon_art_fas, encoding='cp1251').read().splitlines()
        for i in ozon_art_file:
            p = i.split('\t')
            ozon_art.append([p[0], int(p[1])])

        for i in ozon_art:
            art, fas = i[0], i[1]
            flag = False

            name_art = art
            if 'ozon_swap' in locals() and art in ozon_swap.keys():
                art = ozon_swap[art]

            for p in ostatki:
                if p[0] == art[:11]:
                    if type(p[1]) is int:
                        ost = p[1] // fas
                        ost = round(ost * float(ozon_koef)) - predel
                    elif type(p[1]) is float:
                        sg.Print(f'ОЗОН-остатки. Проверьте артикул: {art} - дробное или нулевое значение остатка')
                        continue
                    if ost <= 0:
                        ost = 0
                    ozon_ostatki.append([name_sklad_ozon, name_art, None, ost])
                    flag = True
            if not flag:
                ozon_ostatki.append([name_sklad_ozon, name_art, None, 0])
        try:
            ozon_ostatki_to_df = pd.DataFrame(ozon_ostatki[1:], columns=[ozon_ostatki[0][0], ozon_ostatki[0][1],
                                                                         ozon_ostatki[0][2], ozon_ostatki[0][3]])
            ozon_file = pd.ExcelFile(shab_ozon)
            sheet_to_df_map = {}
            for sheet in ozon_file.sheet_names:
                sheet_to_df_map[sheet] = ozon_file.parse(sheet, index_col=0)
            sheet_to_df_map['Остатки на складе'] = ozon_ostatki_to_df
            sheet_to_df_map['Инструкция'] = pd.DataFrame(['Медтехника 2.0 FBS Армавир (23939175582000)'], columns=None)
            with pd.ExcelWriter(f'{path}/ozon.xlsx', engine='xlsxwriter', mode='w') as excel_writer:
                for sheet_name in ozon_file.sheet_names:
                    sheet_to_df_map[sheet_name].to_excel(excel_writer, sheet_name=sheet_name, index=False)
            sg.Print('Остатки OZON записаны')
        except Exception as ex:
            sg.Print(ex)
        if use_api_flag:
            warehouse_id = int(name_sklad_ozon[-15:-1])
            api_stocks_list = [{'offer_id': i[1], 'stock': i[3], 'warehouse_id': warehouse_id} for i in ozon_ostatki[1:]]

            n = 100 #макисмальное значение пакета для отправки
            while len(api_stocks_list) > n:
                response = api.update_stocks_ozon(api_stocks_list[:n])
                api_stocks_list = api_stocks_list[n:]
                time.sleep(1)
            response = api.update_stocks_ozon(api_stocks_list)


            if response[0] == 200 and len(response[1]) == 0:
                sg.Print('Остатки OZON обновлены')
            else:
                sg.Print(response[0], response[1])

    if ya_art_fas and not sber_flag:
        ya_art_file = open(ya_art_fas, encoding='cp1251').read().splitlines()

        for i in ya_art_file:
            p = i.split('\t')
            ya_art.append([p[0], int(p[1])])

        for i in ya_art:
            art, fas = i[0], i[1]
            flag = False
            name_art = art
            if 'ozon_swap' in locals() and art in ozon_swap.keys():
                art = ozon_swap[art]
            for p in ostatki:
                if p[0] == art[:11]:
                    if type(p[1]) is int:
                        ost = p[1] // fas
                        ost = round(ost * float(ya_koef)) - predel
                    elif type(p[1]) is float:
                        sg.Print(f'Я-остатки. Проверьте артикул: {art} - дробное или нулевое значение остатка')
                        continue
                    if ost <= 0:
                        ost = 0
                    ya_ostatki.append([None, None, name_art, None, ost])
                    flag = True
            if not flag:
                ya_ostatki.append([None, None, name_art, None, 0])

        try:
            ya = pd.ExcelFile(shab_ya)
            sheet_to_df_map = {}
            for sheet in ya.sheet_names:
                sheet_to_df_map[sheet] = ya.parse(sheet, index_col=0)
            ya_ostatki_to_df = pd.DataFrame(ya_ostatki[1:],
                                            columns=[ya_ostatki[0][0], ya_ostatki[0][1], ya_ostatki[0][2],
                                                     ya_ostatki[0][3], ya_ostatki[0][4]])
            sheet_to_df_map['Остатки'] = ya_ostatki_to_df
            sheet_to_df_map['Инструкция'] = pd.DataFrame(None, columns=['ЯФ2244О'])

            with pd.ExcelWriter(f'{path}/ya.xlsx', engine='xlsxwriter', mode='w') as excel_writer:
                for sheet_name in ya.sheet_names:
                    if sheet_name == 'Инструкция':
                        sheet_to_df_map[sheet_name].to_excel(excel_writer, sheet_name=sheet_name, index=False)
                    sheet_to_df_map[sheet_name].to_excel(excel_writer, sheet_name=sheet_name, index=False)
            sg.Print('Остатки Яндекс записаны')
        except Exception as ex:
            sg.Print(ex)
        if use_api_flag:
            token_dict = json.load(open('token_yandex.txt'))
            token_ya = token_dict['Token']
            response_status_code, response = api.update_yandex_stocks(token_ya, ya_ostatki)
            if response_status_code == 200:
                sg.Print('Остатки Яндекс обновлены')
            else:
                sg.Print(f'Остатки Яндекс обновить не удалось\nКод ошибки: {response_status_code}\nОтвет сервера: {response}')

